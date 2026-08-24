"""TASK-020 proof: Round-trip data (Wave 2) -- multi-format ingestion + export.

Standalone and idempotent (AP-7): tears down its own fixtures first, so it runs
twice consecutively with identical results. Announces the live Redis backend (AP-9)
-- if that is not `redis`, the cache-touching proof is void. Must run with the
uvicorn backend STOPPED: it uses the real single-file spencer.db via db_manager,
which holds a single write lock.

Covers the Wave 2 acceptance criteria with real, printed output:
  1. export_service.encode_table(fmt) produces bytes for every format, and each
     round-trips back through the REAL ingestion path (analyze_and_register_table,
     which dispatches by extension) with row count + column names preserved --
     proving #31 (multi-format ingestion) and #10 (table export) together.
  2. Parquet preserves column TYPES exactly (the #31 parquet-bug guard: csv/tsv/json
     re-infer types, parquet does not).
  3. xlsx string cells are stored as text, so a leading '=' is preserved literally
     and can never be read as a spreadsheet formula (injection-safe fidelity).
  4. encode_rows (#24 query-result xlsx) round-trips header + rows, and both entry
     points fail closed with ExportError on an unsupported format (router -> 400).
"""
import asyncio
import os
import sys
from io import BytesIO

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from services.redis_manager import redis_manager
from services.duckdb_manager import db_manager
from services import export_service
from routers.session import analyze_and_register_table, _table_name_for

SRC = "test_exp_src"
RT_FORMATS = ("csv", "tsv", "json", "parquet", "xlsx")

_failures = []


def check(label, cond, detail=""):
    mark = "PASS" if cond else "FAIL"
    if not cond:
        _failures.append(label)
    line = f"  [{mark}] {label}"
    if detail:
        line += f"  ->  {detail}"
    print(line)


async def _rowcount(tbl):
    return (await db_manager.run_readwrite(f'SELECT COUNT(*) FROM "{tbl}"'))[0][0]


async def _cols(tbl):
    info = await db_manager.run_readwrite(f'PRAGMA table_info("{tbl}")')
    return [row[1] for row in info]


async def _types(tbl):
    info = await db_manager.run_readwrite(f'PRAGMA table_info("{tbl}")')
    return {row[1]: row[2] for row in info}


async def _drop_like(pattern):
    rows = await db_manager.run_readwrite(
        "SELECT table_name FROM information_schema.tables "
        "WHERE table_schema = 'main' AND table_name LIKE ?",
        (pattern,),
    )
    for (tn,) in rows or []:
        await db_manager.run_readwrite(f'DROP TABLE IF EXISTS "{tn}"')


def _redis_cleanup(*sessions):
    c = redis_manager.client
    for s in sessions:
        c.delete(f"schema:{s}")
        c.delete(f"schema_version:{s}")
        for k in c.keys(f"history:{s}:*"):
            c.delete(k)


async def _teardown():
    await _drop_like("%test_exp%")
    _redis_cleanup(*[f"test-exp-rt-{f}" for f in RT_FORMATS])


def _tmp_write(name, data):
    os.makedirs("uploads/_tasktest", exist_ok=True)
    path = f"uploads/_tasktest/{name}"
    with open(path, "wb") as f:
        f.write(data)
    return path


async def _seed_source():
    """Direct typed CREATE so parquet fidelity + tricky values are fully controlled:
    a leading-'=' formula-injection string, a value containing a comma/semicolon/tab
    (delimiter robustness), unicode, and an all-NULL row."""
    await db_manager.run_readwrite(f'DROP TABLE IF EXISTS "{SRC}"')
    await db_manager.run_readwrite(
        f'''CREATE TABLE "{SRC}" AS SELECT * FROM (VALUES
            (1, 'plain',        12.5,  DATE '2024-01-15', TRUE),
            (2, '=SUM(A1)',     -3.0,  DATE '2024-02-20', FALSE),
            (3, 'a,b;c\tx',      0.0,   DATE '2024-03-10', TRUE),
            (4, 'café ünïcode', 99.99, DATE '2024-06-30', FALSE),
            (5, NULL,           NULL,  NULL,              NULL)
        ) AS t(id, label, amount, d, flag)'''
    )


async def service_main():
    print("=" * 70)
    print("TASK-020 PROOF -- round-trip data (multi-format ingestion + export)")
    print(f"REDIS BACKEND IN USE: {redis_manager.backend}")
    if redis_manager.server_version:
        print(f"  (real redis-server version {redis_manager.server_version})")
    print("=" * 70)

    await _teardown()
    await _seed_source()
    src_rows = await _rowcount(SRC)
    src_cols = await _cols(SRC)
    src_types = await _types(SRC)
    print(f"\nSeeded {SRC}: {src_rows} rows, cols={src_cols}")

    # --- 1. Export each format, then re-ingest through the real ingestion path ---
    print("\n--- 1. Export each format, then re-ingest (round-trip) ---")
    for fmt in export_service.TABLE_FORMATS:
        data = await export_service.encode_table(SRC, fmt)
        check(
            f"{fmt}: encode_table returns non-empty bytes",
            isinstance(data, (bytes, bytearray)) and len(data) > 0,
            f"{len(data)} bytes",
        )
        path = _tmp_write(f"rt.{fmt}", data)
        sess = f"test-exp-rt-{fmt}"
        tname, _ = _table_name_for(sess, f"rt.{fmt}")
        await analyze_and_register_table(sess, tname, path, is_primary=True)
        rc = await _rowcount(tname)
        cols = await _cols(tname)
        check(f"{fmt}: round-trip row count preserved ({src_rows})", rc == src_rows, f"got {rc}")
        check(f"{fmt}: round-trip column names preserved", cols == src_cols, f"got {cols}")

    # --- 2. Parquet preserves column types exactly (the #31 parquet-bug guard) ---
    print("\n--- 2. Parquet preserves column types exactly ---")
    ptname, _ = _table_name_for("test-exp-rt-parquet", "rt.parquet")
    ptypes = await _types(ptname)
    check("parquet: types identical to source", ptypes == src_types, f"{ptypes} vs {src_types}")

    # --- 3. xlsx string cells stored as text (formula-injection safe) ---
    print("\n--- 3. xlsx string cells stored as text (formula-injection safe) ---")
    from openpyxl import load_workbook

    xdata = await export_service.encode_table(SRC, "xlsx")
    wb = load_workbook(BytesIO(xdata), read_only=True)
    grid = [[c.value for c in row] for row in wb.active.iter_rows()]
    wb.close()
    check("xlsx: header row = column names", grid[0] == src_cols, f"{grid[0]}")
    # openpyxl drops trailing empty cells, so the all-NULL row reads back short --
    # guard the index (this same trimming is what the xlsx-bridge fix null-pads).
    labels = [r[1] if len(r) > 1 else None for r in grid[1:]]
    check("xlsx: leading-'=' value preserved literally as text", "=SUM(A1)" in labels, f"{labels}")

    # --- 4. encode_rows (#24 query-result xlsx) + format guards ---
    print("\n--- 4. encode_rows (query-result xlsx) + format guards ---")
    rows = [{"a": 1, "b": "=cmd"}, {"a": 2, "b": "x"}]
    rdata = await export_service.encode_rows(["a", "b"], rows, "xlsx")
    wb = load_workbook(BytesIO(rdata), read_only=True)
    rgrid = [[c.value for c in row] for row in wb.active.iter_rows()]
    wb.close()
    check("encode_rows: header + 2 data rows", rgrid[0] == ["a", "b"] and len(rgrid) == 3, f"{rgrid}")
    check("encode_rows: leading-'=' cell preserved literally", rgrid[1][1] == "=cmd", f"{rgrid[1]}")

    rows_guard = False
    try:
        await export_service.encode_rows(["a"], [{"a": 1}], "csv")
    except export_service.ExportError:
        rows_guard = True
    check("encode_rows rejects non-xlsx (csv) with ExportError", rows_guard)

    table_guard = False
    try:
        await export_service.encode_table(SRC, "xml")
    except export_service.ExportError:
        table_guard = True
    check("encode_table rejects unsupported format (xml) with ExportError", table_guard)

    await _teardown()

    print("\n" + "=" * 70)
    if _failures:
        print(f"RESULT: {len(_failures)} FAILURE(S): {_failures}")
    else:
        print("RESULT: ALL CHECKS PASSED")
    print("=" * 70)
    return 1 if _failures else 0


if __name__ == "__main__":
    sys.exit(asyncio.run(service_main()))
