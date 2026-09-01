import uuid
import os
import re
import csv
from typing import Optional
from fastapi import APIRouter, UploadFile, File, HTTPException, Depends
from sqlalchemy.orm import Session
from starlette.concurrency import run_in_threadpool
import config
from models.schemas import (
    SessionResponse,
    SchemaResponse,
    TableUploadResponse,
    MaterializeRequest,
    TableSchemaResponse,
    TransformParam,
    TransformResponse,
    TransformPreviewResponse,
    HistoryResponse,
    ColumnSchema
)
from deps import get_current_user, get_db, require_session_owner
from services.app_db import Dataset, User
from services.duckdb_manager import db_manager, current_session_id
from services.redis_manager import redis_manager
from services.sql_validator import sql_validator
from services import cleanup_service, ownership_service, transform_service

router = APIRouter()

def sanitize_table_name(name: str) -> str:
    clean = re.sub(r'\W', '_', name)
    clean = re.sub(r'^_+|_+$', '', clean)
    if not clean or clean[0].isdigit():
        clean = "t_" + clean
    return clean.lower()

def _safe_disk_name(clean_name: str, original_filename: str) -> str:
    """A filesystem-safe upload name. `clean_name` is already \\W-sanitized; we
    keep only the original extension and strip anything that could enable path
    traversal or SQL-string breakout from the extension."""
    ext = os.path.splitext(original_filename)[1]
    return re.sub(r'[^A-Za-z0-9._-]', '_', f"{clean_name}{ext}")

def _table_name_for(session_uuid: str, filename: str):
    base_name = os.path.splitext(filename)[0]
    clean_name = sanitize_table_name(base_name)
    return f"t_{session_uuid.replace('-', '_')}_{clean_name}", clean_name

def _reject_disallowed_type(filename: str) -> None:
    """415 if the upload's extension is not in the configured allowlist. Fails
    closed (a name with no extension is rejected) and runs before any bytes are
    persisted -- today a non-CSV instead fails deep inside read_csv_auto."""
    if not config.is_allowed_upload(filename):
        allowed = ", ".join(sorted(config.ALLOWED_EXTENSIONS)) or "(none)"
        raise HTTPException(
            status_code=415,
            detail=f"Unsupported file type. Allowed extensions: {allowed}",
        )

def _persist_upload(session_uuid: str, file: UploadFile, clean_name: str) -> str:
    """Stream the upload to disk in fixed chunks, counting bytes. On exceeding
    the cap, delete the partial file and raise 413. This is the streaming
    backstop layer -- it catches chunked / absent / lying Content-Length that
    the main.py middleware's honest-Content-Length check cannot."""
    os.makedirs(f"{config.UPLOADS_DIR}/{session_uuid}", exist_ok=True)
    file_path = f"{config.UPLOADS_DIR}/{session_uuid}/{_safe_disk_name(clean_name, file.filename)}"
    written = 0
    try:
        with open(file_path, "wb") as buffer:
            while True:
                chunk = file.file.read(config.UPLOAD_CHUNK_BYTES)
                if not chunk:
                    break
                written += len(chunk)
                if written > config.MAX_UPLOAD_BYTES:
                    raise HTTPException(
                        status_code=413,
                        detail=f"Upload exceeds the {config.MAX_UPLOAD_MB} MB limit",
                    )
                buffer.write(chunk)
    except HTTPException:
        # Leave no residue: remove only the partial file (not the session dir,
        # which for a second-table upload holds the first table's file).
        if os.path.exists(file_path):
            os.remove(file_path)
        raise
    return file_path

def _quote_ident(name: str) -> str:
    """Double-quote a SQL identifier, escaping embedded quotes (column names
    come from CSV headers, which are not trusted)."""
    return '"' + name.replace('"', '""') + '"'

async def refresh_table_schema_cache(session_uuid: str, table_name: str, is_primary: bool):
    """Compute cardinality/samples/DDL for `table_name` and (re)write the
    `schema:{session}` cache entry; return the ColumnSchema list.

    Shared by ingestion and by post-transform refresh: after any transform the
    schema is recomputed from the live table rather than trusted from cache
    (ARCHITECTURE.md — schema is never statically cached across a transform)."""
    col_info = await db_manager.run_readwrite(f"PRAGMA table_info({table_name})")
    columns = []

    table_schema_context = {
        "cardinality": {},
        "samples": {},
        "is_primary": is_primary
    }

    ddl_parts = []

    for row in col_info:
        col_name = row[1]
        col_type = row[2]
        qcol = _quote_ident(col_name)

        card_res = await db_manager.run_readwrite(f'SELECT COUNT(DISTINCT {qcol}) FROM {table_name}')
        cardinality = card_res[0][0] if card_res else 0

        columns.append(ColumnSchema(name=col_name, type=col_type, cardinality=cardinality))

        table_schema_context["cardinality"][col_name] = cardinality
        ddl_parts.append(f'{qcol} {col_type}')

        if cardinality < 20 and cardinality > 0:
            samples_res = await db_manager.run_readwrite(
                f'SELECT DISTINCT {qcol} FROM {table_name} WHERE {qcol} IS NOT NULL LIMIT 20'
            )
            samples = [s[0] for s in samples_res]
            table_schema_context["samples"][col_name] = samples

    table_schema_context["ddl"] = f"CREATE TABLE {table_name} ({', '.join(ddl_parts)});"

    schema_key = f"schema:{session_uuid}"
    schema_data = redis_manager.get_json(schema_key) or {}
    schema_data[table_name] = table_schema_context
    redis_manager.set_json(schema_key, schema_data)

    return columns


def _reader_sql(ext: str) -> Optional[str]:
    """DuckDB table-function for a natively-readable upload, file path BOUND as `?`
    (never interpolated). Returns None for formats needing a Python bridge (xlsx)."""
    if ext == "csv":
        return "read_csv_auto(?, header=true)"
    if ext == "tsv":
        return "read_csv_auto(?, header=true, delim='\t')"
    if ext == "parquet":
        return "read_parquet(?)"
    if ext == "json":
        return "read_json_auto(?)"
    return None


def _xlsx_to_csv(xlsx_path: str) -> str:
    """Convert an .xlsx's active sheet to a sibling CSV so DuckDB's read_csv_auto
    (and its type inference) ingests it on the same bound-param path as every other
    format. Blocking (openpyxl + IO) -> callers run it in a threadpool. Returns the
    temp CSV path for the caller to read then delete."""
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        ws = wb.active
        if ws is None:
            raise HTTPException(status_code=400, detail="Spreadsheet has no readable sheet")
        csv_path = xlsx_path + ".converted.csv"
        wrote_any = False
        with open(csv_path, "w", newline="", encoding="utf-8") as fh:
            writer = csv.writer(fh)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(["" if v is None else v for v in row])
                wrote_any = True
        if not wrote_any:
            if os.path.exists(csv_path):
                os.remove(csv_path)
            raise HTTPException(status_code=400, detail="Spreadsheet is empty")
        return csv_path
    finally:
        wb.close()


async def analyze_and_register_table(session_uuid: str, table_name: str, file_path: str, is_primary: bool):
    # Route by extension to the right DuckDB reader (Foundation 3, Wave 2). The file
    # path is BOUND as a parameter (never interpolated) so a crafted filename cannot
    # break out of the reader's string literal; table_name is a \\W-sanitized
    # identifier, safe to interpolate.
    ext = config.ext_of(file_path)
    reader = _reader_sql(ext)
    tmp_csv: Optional[str] = None
    try:
        if reader is None and ext == "xlsx":
            # No native DuckDB reader: bridge the sheet to a temp CSV (openpyxl),
            # then read it on the identical bound-param path as every other format.
            # We WROTE this CSV ourselves (csv.writer -> comma, '"'-quoted), so pin
            # those params instead of auto-sniffing: openpyxl drops trailing empty
            # cells, so a sheet whose last rows don't fill every column yields a
            # ragged CSV that would derail the delimiter sniffer into one giant
            # column. Pinning the delimiter + null_padding pads those short rows
            # with NULLs, so trailing blanks in Excel ingest correctly.
            tmp_csv = await run_in_threadpool(_xlsx_to_csv, file_path)
            read_path, reader = tmp_csv, (
                "read_csv_auto(?, header=true, delim=',', quote='\"', "
                "escape='\"', null_padding=true)"
            )
        else:
            read_path = file_path
        if reader is None:
            # Defense in depth: _reject_disallowed_type should already have 415'd
            # anything we cannot read.
            raise HTTPException(status_code=415, detail=f"Unsupported file type: .{ext}")
        await db_manager.run_readwrite(
            f"CREATE TABLE {table_name} AS SELECT * FROM {reader}",
            (read_path,),
        )
    finally:
        if tmp_csv and os.path.exists(tmp_csv):
            os.remove(tmp_csv)

    row_count = (await db_manager.run_readwrite(f"SELECT COUNT(*) FROM {table_name}"))[0][0]
    columns = await refresh_table_schema_cache(session_uuid, table_name, is_primary)

    return row_count, columns


def _resolve_table(session_uuid: str, table_name: Optional[str]) -> str:
    """Resolve which table an op targets: the named one if it exists in this
    session, otherwise the session's primary table. 404 if the session has no
    tables or the named table is unknown."""
    schema = redis_manager.get_json(f"schema:{session_uuid}") or {}
    if not schema:
        raise HTTPException(status_code=404, detail="No tables in this session")
    if table_name:
        if table_name in schema:
            return table_name
        raise HTTPException(status_code=404, detail=f"Table '{table_name}' not found in session")
    for tname, ctx in schema.items():
        if ctx.get("is_primary"):
            return tname
    return next(iter(schema))

@router.post("", response_model=SessionResponse)
async def create_session(
    file: UploadFile = File(...),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    session_uuid = str(uuid.uuid4())
    current_session_id.set(session_uuid)
    _reject_disallowed_type(file.filename)
    table_name, clean_name = _table_name_for(session_uuid, file.filename)
    # Mark the session live BEFORE persisting so the marker always exists before
    # the upload dir -- the sweeper can never race a just-created session. An
    # owned session gets the longer owned-session TTL (created only by an authed
    # user now), superseding the anonymous 24h default.
    redis_manager.touch_session(session_uuid, config.OWNED_SESSION_TTL_SECONDS)
    file_path = _persist_upload(session_uuid, file, clean_name)

    row_count, columns = await analyze_and_register_table(session_uuid, table_name, file_path, is_primary=True)

    # Record ownership AFTER the data exists, so this user (and only this user)
    # can reach the session. A failure here leaves an orphan table the sweeper
    # reclaims once the marker lapses -- acceptable, and it never leaks to anyone.
    ownership_service.record_dataset(db, session_uuid, user.id, table_name, file.filename)

    return SessionResponse(
        session_uuid=session_uuid,
        table_name=table_name,
        row_count=row_count,
        columns=columns
    )

@router.post("/{session_uuid}/tables", response_model=TableUploadResponse, dependencies=[Depends(require_session_owner)])
async def upload_table(session_uuid: str, file: UploadFile = File(...)):
    _reject_disallowed_type(file.filename)
    table_name, clean_name = _table_name_for(session_uuid, file.filename)

    # Refuse to overwrite an existing table in this session (no silent clobber).
    existing = redis_manager.get_json(f"schema:{session_uuid}") or {}
    if table_name in existing:
        raise HTTPException(
            status_code=409,
            detail=f"Table '{table_name}' already exists in this session"
        )

    # Refresh the liveness marker before persisting (anti-race, as in create).
    # Owned TTL so a second-table upload doesn't downgrade the session's lifetime.
    redis_manager.touch_session(session_uuid, config.OWNED_SESSION_TTL_SECONDS)
    file_path = _persist_upload(session_uuid, file, clean_name)

    row_count, columns = await analyze_and_register_table(session_uuid, table_name, file_path, is_primary=False)

    return TableUploadResponse(
        table_name=table_name,
        row_count=row_count,
        columns=columns
    )

@router.post("/{session_uuid}/materialize", response_model=TableUploadResponse, dependencies=[Depends(require_session_owner)])
async def materialize_query(session_uuid: str, payload: MaterializeRequest):
    """#23: persist a reviewed Query Engine SELECT as a real session table, so a result
    can become a new working table (Table view) or the base for a Canvas chart.

    Same fail-closed AI-SQL defense as /execute (ai.py) -- validate() (a single
    read-only SELECT) then scope_violation() (S-1/TASK-029: only THIS session's own
    tables, no filesystem/IO functions, no qualified names). The ONLY difference is
    that /execute runs in a rolled-back sandbox while this runs on run_readwrite so the
    CREATE TABLE persists. The new table is named + registered exactly like an upload
    (schema cache refreshed, is_primary=False, 409 on a name clash) and its own
    `t_<uuid>_...` name is scope-valid, so it can be queried again afterwards."""
    sql = (payload.sql or "").strip()
    if not sql:
        raise HTTPException(status_code=400, detail="SQL is required.")
    if not sql_validator.validate(sql):
        raise HTTPException(
            status_code=400,
            detail="This query was rejected: only a single read-only SELECT is allowed.",
        )
    # S-1: read-only is necessary but not sufficient on the shared single-file DuckDB --
    # gate the SELECT to this session's own tables (identical check to /execute).
    scope_reason = sql_validator.scope_violation(sql, session_uuid)
    if scope_reason:
        raise HTTPException(
            status_code=400,
            detail=f"This query was rejected: it {scope_reason}.",
        )

    clean_name = sanitize_table_name(payload.name or "query_result")
    table_name = f"t_{session_uuid.replace('-', '_')}_{clean_name}"

    # No silent clobber -- same guard as upload_table.
    existing = redis_manager.get_json(f"schema:{session_uuid}") or {}
    if table_name in existing:
        raise HTTPException(
            status_code=409,
            detail=f"A table named '{clean_name}' already exists in this session",
        )

    # Refresh the liveness marker (owned TTL), as create/upload do.
    redis_manager.touch_session(session_uuid, config.OWNED_SESSION_TTL_SECONDS)

    # Strip a lone trailing ';' so the SELECT wraps as a subquery (stacked statements
    # were already rejected by validate()).
    inner = sql.rstrip(";").rstrip()
    try:
        # run_readwrite (NOT run_sandboxed): the whole point is to PERSIST the result.
        await db_manager.run_readwrite(f"CREATE TABLE {table_name} AS SELECT * FROM ({inner}) _q")
    except Exception as exc:
        # A validated SELECT can still fail at run time (unknown column, duplicate
        # output column names, type error). Surface it as a 400 for the editor.
        raise HTTPException(status_code=400, detail=f"Could not save result as a table: {exc}")

    row_count = (await db_manager.run_readwrite(f"SELECT COUNT(*) FROM {table_name}"))[0][0]
    columns = await refresh_table_schema_cache(session_uuid, table_name, is_primary=False)

    return TableUploadResponse(
        table_name=table_name,
        row_count=row_count,
        columns=columns
    )

@router.get("/{session_uuid}/schema", response_model=SchemaResponse)
async def get_schema(
    session_uuid: str,
    owner = Depends(require_session_owner),
):
    schema_key = f"schema:{session_uuid}"
    schema_data = redis_manager.get_json(schema_key)
    
    if not schema_data:
        # CRASH RECOVERY: Rebuild from DuckDB
        session_prefix = f"t_{session_uuid.replace('-', '_')}"
        tables_info = await db_manager.run_readwrite(
            f"SELECT table_name FROM information_schema.tables WHERE table_name LIKE '{session_prefix}%'"
        )
        if not tables_info:
            raise HTTPException(status_code=404, detail="Schema not found")
            
        for (tname,) in tables_info:
            is_primary = (tname == owner.primary_table)
            await refresh_table_schema_cache(session_uuid, tname, is_primary)
            
        schema_data = redis_manager.get_json(schema_key)
        if not schema_data:
            raise HTTPException(status_code=404, detail="Schema rebuild failed")

    tables = []
    for t_name, t_data in schema_data.items():
        col_info = await db_manager.run_readwrite(f"PRAGMA table_info({t_name})")
        cols = []
        for row in col_info:
            c_name = row[1]
            c_type = row[2]
            c_card = t_data["cardinality"].get(c_name, 0)
            cols.append(ColumnSchema(name=c_name, type=c_type, cardinality=c_card))

        tables.append(TableSchemaResponse(
            table_name=t_name,
            is_primary=t_data.get("is_primary", False),
            columns=cols
        ))

    return SchemaResponse(tables=tables)

@router.delete("/{session_uuid}/tables/{table_name}", dependencies=[Depends(require_session_owner)])
async def delete_table(session_uuid: str, table_name: str):
    """Drop a table from the session and remove it from the schema cache."""
    # Ensure the table name is valid/safe before using in a raw DROP
    import re
    if not re.match(r"^[a-zA-Z0-9_]+$", table_name):
        raise HTTPException(status_code=400, detail="Invalid table name")

    schema_key = f"schema:{session_uuid}"
    schema_data = redis_manager.get_json(schema_key)
    if not schema_data or table_name not in schema_data:
        raise HTTPException(status_code=404, detail=f"Table '{table_name}' not found in session")

    # Prevent dropping the primary table if it's the only one (optional safety, or just let them)
    # Actually, if they drop the primary table, they can just upload a new one.

    # Drop from DuckDB
    try:
        await db_manager.run_readwrite(f"DROP TABLE IF EXISTS {table_name}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

    # Remove from Redis schema cache
    del schema_data[table_name]
    redis_manager.set_json(schema_key, schema_data)

    return {"status": "ok"}

@router.post("/{session_uuid}/tables/{table_name}/primary", dependencies=[Depends(require_session_owner)])
async def make_table_primary(session_uuid: str, table_name: str):
    """Set a specific table as the primary table for the session."""
    schema_key = f"schema:{session_uuid}"
    schema_data = redis_manager.get_json(schema_key)
    if not schema_data or table_name not in schema_data:
        raise HTTPException(status_code=404, detail=f"Table '{table_name}' not found in session")

    # Update primary flag
    for tname, ctx in schema_data.items():
        ctx["is_primary"] = (tname == table_name)
    
    redis_manager.set_json(schema_key, schema_data)
    return {"status": "ok"}

@router.delete("/{session_uuid}")
async def delete_session(
    session_uuid: str,
    _owner: Dataset = Depends(require_session_owner),
    db: Session = Depends(get_db),
):
    """Owner-initiated teardown: drop the session's DuckDB tables, purge its Redis
    keys, remove its upload dir, and delete the ownership row. The ownership guard
    404s if the session is already gone or not the caller's, so this is safe to
    call once; a second call 404s rather than double-freeing."""
    counts = await cleanup_service.reclaim_session_storage(session_uuid)
    ownership_service.delete_dataset(db, session_uuid)
    return {"status": "deleted", **counts}

@router.post("/{session_uuid}/transform", response_model=TransformResponse, dependencies=[Depends(require_session_owner)])
async def apply_transform(session_uuid: str, payload: TransformParam, table_name: Optional[str] = None):
    tname = _resolve_table(session_uuid, table_name)
    try:
        version, step, row_count = await transform_service.apply_transform(session_uuid, tname, payload)
    except transform_service.TransformError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    # Recompute the schema cache from the transformed table (types/cardinality
    # may have changed). Preserve the table's primary flag.
    schema = redis_manager.get_json(f"schema:{session_uuid}") or {}
    is_primary = schema.get(tname, {}).get("is_primary", False)
    await refresh_table_schema_cache(session_uuid, tname, is_primary)
    return TransformResponse(schema_version=version, step=step, row_count=row_count)

@router.post("/{session_uuid}/transform/preview", response_model=TransformPreviewResponse, dependencies=[Depends(require_session_owner)])
async def preview_transform(session_uuid: str, payload: TransformParam, table_name: Optional[str] = None):
    """Dry-run: report what the op WOULD do (row-count delta, resulting schema, a
    sample) without applying it -- no history step, no schema_version bump. Same
    fail-closed validation as the real transform."""
    tname = _resolve_table(session_uuid, table_name)
    try:
        result = await transform_service.preview_transform(session_uuid, tname, payload)
    except transform_service.TransformError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    return TransformPreviewResponse(**result)

@router.post("/{session_uuid}/undo", response_model=TransformResponse, dependencies=[Depends(require_session_owner)])
async def undo_transform(session_uuid: str, table_name: Optional[str] = None):
    tname = _resolve_table(session_uuid, table_name)
    try:
        version, step, row_count = await transform_service.undo(session_uuid, tname)
    except transform_service.TransformError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    schema = redis_manager.get_json(f"schema:{session_uuid}") or {}
    is_primary = schema.get(tname, {}).get("is_primary", False)
    await refresh_table_schema_cache(session_uuid, tname, is_primary)
    return TransformResponse(schema_version=version, step=step, row_count=row_count)

@router.post("/{session_uuid}/redo", response_model=TransformResponse, dependencies=[Depends(require_session_owner)])
async def redo_transform(session_uuid: str, table_name: Optional[str] = None):
    tname = _resolve_table(session_uuid, table_name)
    try:
        version, step, row_count = await transform_service.redo(session_uuid, tname)
    except transform_service.TransformError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    schema = redis_manager.get_json(f"schema:{session_uuid}") or {}
    is_primary = schema.get(tname, {}).get("is_primary", False)
    await refresh_table_schema_cache(session_uuid, tname, is_primary)
    return TransformResponse(schema_version=version, step=step, row_count=row_count)

@router.get("/{session_uuid}/history", response_model=HistoryResponse, dependencies=[Depends(require_session_owner)])
async def get_history(session_uuid: str, table_name: Optional[str] = None):
    tname = _resolve_table(session_uuid, table_name)
    return HistoryResponse(**transform_service.get_history(session_uuid, tname))
